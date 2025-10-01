"""
Main Processor Module
Orchestrates PDF processing, sentence splitting, and output generation
"""

import os
import time
import logging
from datetime import datetime
from typing import List, Optional, Callable, Dict, Any
import pandas as pd
from PyPDF2 import PdfReader
from pdf2image import convert_from_path
import pytesseract
from src.core.sentence_splitter import SentenceSplitter, ProcessingMode, SentenceResult
from src.utils.config_manager import ConfigManager
from src.utils.google_sheets import GoogleSheetsManager

logger = logging.getLogger(__name__)


class NovelProcessor:
    """Main processor for French novels"""
    
    def __init__(self, config_manager: ConfigManager):
        """
        Initialize processor
        
        Args:
            config_manager: Configuration manager instance
        """
        self.config = config_manager
        self.results: List[SentenceResult] = []
        self.processing_time = 0
        self.start_time = None
        # Reference to the active SentenceSplitter while processing
        self._active_splitter = None
    
    def extract_text_from_pdf(self, pdf_path: str, progress_callback: Optional[Callable] = None) -> str:
        """
        Extract text from PDF file
        
        Args:
            pdf_path: Path to PDF file
            progress_callback: Optional callback for progress updates
            
        Returns:
            Extracted text
        """
        text = ""
        
        try:
            # Try direct text extraction first
            reader = PdfReader(pdf_path)
            num_pages = len(reader.pages)
            
            for i, page in enumerate(reader.pages):
                if progress_callback:
                    progress_callback(i + 1, num_pages, "Extracting text from PDF...")
                
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            
            # If text extraction yielded little content, try OCR
            if len(text.strip()) < 100:
                if progress_callback:
                    progress_callback(0, num_pages, "Text extraction insufficient, trying OCR...")
                
                text = self.extract_text_with_ocr(pdf_path, progress_callback)
        
        except Exception as e:
            logger.error(f"Error extracting text: {str(e)}")
            # Try OCR as fallback
            if progress_callback:
                progress_callback(0, 0, "Falling back to OCR...")
            text = self.extract_text_with_ocr(pdf_path, progress_callback)
        
        return text
    
    def extract_text_with_ocr(self, pdf_path: str, progress_callback: Optional[Callable] = None) -> str:
        """
        Extract text using OCR
        
        Args:
            pdf_path: Path to PDF file
            progress_callback: Optional callback for progress updates
            
        Returns:
            Extracted text
        """
        text = ""
        
        try:
            # Convert PDF to images
            images = convert_from_path(pdf_path)
            
            for i, image in enumerate(images):
                if progress_callback:
                    progress_callback(i + 1, len(images), "Running OCR...")
                
                # Run OCR with French language
                page_text = pytesseract.image_to_string(image, lang='fra')
                text += page_text + "\n"
        
        except Exception as e:
            logger.error(f"OCR error: {str(e)}")
            raise Exception(f"Failed to extract text from PDF: {str(e)}")
        
        return text
    
    def process_pdf(self, pdf_path: str, word_limit: Optional[int] = None, 
                   processing_mode: Optional[str] = None,
                   progress_callback: Optional[Callable] = None) -> List[SentenceResult]:
        """
        Process a PDF file
        
        Args:
            pdf_path: Path to PDF file
            word_limit: Word limit per sentence (uses config default if None)
            processing_mode: Processing mode (uses config default if None)
            progress_callback: Optional callback for progress updates
            
        Returns:
            List of SentenceResult objects
        """
        self.start_time = time.time()
        
        # Get settings
        if word_limit is None:
            word_limit = self.config.get_word_limit()
        
        if processing_mode is None:
            processing_mode = self.config.get_processing_mode()
        
        # Extract text from PDF
        if progress_callback:
            progress_callback(0, 100, "Extracting text from PDF...")
        
        text = self.extract_text_from_pdf(pdf_path, progress_callback)
        
        if not text or len(text.strip()) < 10:
            raise Exception("No text extracted from PDF")
        
        # Initialize sentence splitter
        mode = ProcessingMode.AI_REWRITE if processing_mode == 'ai_rewrite' else ProcessingMode.MECHANICAL_CHUNKING
        
        # Smart API selection: Gemini (dev) or OpenAI (production)
        api_key = None
        use_gemini = False
        if mode == ProcessingMode.AI_REWRITE:
            if self.config.should_use_gemini():
                api_key = self.config.get_gemini_api_key()
                use_gemini = True
                if progress_callback:
                    progress_callback(45, 100, "Using Gemini AI (Development Mode)...")
            else:
                api_key = self.config.get_api_key()
                model = self.config.get_openai_model()
                if progress_callback:
                    progress_callback(45, 100, f"Using OpenAI {model}...")

            if not api_key:
                raise Exception("API key required for AI rewriting mode. Please configure it in settings.")

        # Create splitter for the chosen mode (AI or mechanical)
        splitter = SentenceSplitter(word_limit=word_limit, mode=mode, api_key=api_key, use_gemini=use_gemini)

        # Process sentences
        if progress_callback:
            progress_callback(50, 100, "Processing sentences...")

        # Bind our results list to the splitter's live results so get_summary() can observe
        # incremental updates while processing.
        self.results = splitter.results
        results = splitter.process_text(text, progress_callback)
        # Ensure final results are the splitter's results
        self.results = results

        self.processing_time = time.time() - self.start_time

        return self.results
    
    def generate_dataframe(self) -> pd.DataFrame:
        """
        Generate pandas DataFrame from results
        
        Returns:
            DataFrame with processed sentences
        """
        rows = []
        row_num = 1
        
        show_original = self.config.get_show_original()
        
        for result in self.results:
            for sentence in result.output_sentences:
                row = {
                    'Row': row_num,
                    'Sentence': sentence,
                    'Word_Count': len(sentence.split())
                }
                
                if show_original and result.method != "Direct":
                    row['Original'] = result.original
                    row['Method'] = result.method
                else:
                    row['Original'] = ''
                    row['Method'] = result.method if result.method != "Direct" else ''
                
                rows.append(row)
                row_num += 1
        
        # Create DataFrame
        columns = ['Row', 'Sentence']
        if show_original:
            columns.extend(['Original', 'Method'])
        columns.append('Word_Count')
        
        df = pd.DataFrame(rows)
        
        # Reorder columns
        existing_cols = [col for col in columns if col in df.columns]
        df = df[existing_cols]
        
        return df
    
    def generate_processing_log(self) -> pd.DataFrame:
        """
        Generate processing log DataFrame
        
        Returns:
            DataFrame with processing details
        """
        log_rows = []
        
        for i, result in enumerate(self.results):
            if result.method != "Direct":
                log_rows.append({
                    'Sentence_Number': i + 1,
                    'Original': result.original,
                    'Original_Word_Count': result.word_count,
                    'Method': result.method,
                    'Output_Sentences': ' | '.join(result.output_sentences),
                    'Success': result.success,
                    'Error': result.error or ''
                })
        
        return pd.DataFrame(log_rows)
    
    def save_to_csv(self, output_path: str):
        """
        Save results to CSV file
        
        Args:
            output_path: Path to output CSV file
        """
        df = self.generate_dataframe()
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        
        # Save processing log if enabled
        if self.config.get_generate_log():
            log_df = self.generate_processing_log()
            if not log_df.empty:
                log_path = output_path.replace('.csv', '_log.csv')
                log_df.to_csv(log_path, index=False, encoding='utf-8-sig')
    
    def save_to_excel(self, output_path: str):
        """
        Save results to beautifully formatted Excel file
        
        Args:
            output_path: Path to output Excel file
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        df = self.generate_dataframe()
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sentences', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Sentences']
            
            # Define styles
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0')
            )
            
            # Format header row
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Format data rows
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = cell_alignment
                    cell.border = border
                    
                    # Color code by method
                    method = worksheet.cell(row=row, column=df.columns.get_loc('Method') + 1).value
                    if method:
                        if 'AI-Rewritten' in method:
                            cell.fill = PatternFill(start_color='E7F5E7', end_color='E7F5E7', fill_type='solid')
                        elif 'Mechanical' in method:
                            cell.fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')
            
            # Set column widths
            column_widths = {
                'Row': 8,
                'Sentence': 60,
                'Original': 60,
                'Method': 20,
                'Word_Count': 12
            }
            
            for col_num, column in enumerate(df.columns, 1):
                column_letter = get_column_letter(col_num)
                if column in column_widths:
                    worksheet.column_dimensions[column_letter].width = column_widths[column]
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
            # Add processing log if enabled
            if self.config.get_generate_log():
                log_df = self.generate_processing_log()
                if not log_df.empty:
                    log_df.to_excel(writer, sheet_name='Processing Log', index=False)
                    
                    # Format log sheet
                    log_sheet = writer.sheets['Processing Log']
                    for col in range(1, len(log_df.columns) + 1):
                        cell = log_sheet.cell(row=1, column=col)
                        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF', size=11)
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    for col_num in range(1, len(log_df.columns) + 1):
                        column_letter = get_column_letter(col_num)
                        log_sheet.column_dimensions[column_letter].width = 30
                    
                    log_sheet.freeze_panes = 'A2'
            
            # Add summary sheet
            summary = self.get_summary()
            summary_data = {
                'Metric': [
                    'Total Input Sentences',
                    'Total Output Sentences',
                    'Direct (No Processing)',
                    'AI Rewritten',
                    'Mechanical Chunked',
                    'Processing Time',
                    'Average Words per Sentence',
                    'Success Rate'
                ],
                'Value': [
                    summary['total_input_sentences'],
                    summary['total_output_sentences'],
                    summary['direct_sentences'],
                    summary['ai_rewritten'],
                    summary['mechanical_chunked'],
                    f"{summary.get('processing_time', 0):.2f}s",
                    f"{summary['total_output_sentences'] / summary['total_input_sentences']:.2f}" if summary['total_input_sentences'] > 0 else 'N/A',
                    f"{((summary['total_input_sentences'] - summary.get('failed', 0)) / summary['total_input_sentences'] * 100):.1f}%" if summary['total_input_sentences'] > 0 else 'N/A'
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary sheet
            summary_sheet = writer.sheets['Summary']
            for col in range(1, 3):
                cell = summary_sheet.cell(row=1, column=col)
                cell.fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF', size=11)
                cell.alignment = header_alignment
                cell.border = border
            
            summary_sheet.column_dimensions['A'].width = 30
            summary_sheet.column_dimensions['B'].width = 20
            
            for row in range(2, len(summary_df) + 2):
                for col in range(1, 3):
                    cell = summary_sheet.cell(row=row, column=col)
                    cell.border = border
                    if col == 1:
                        cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            summary_sheet.freeze_panes = 'A2'
    
    def save_to_google_sheets(self, title: str, credentials_path: str = 'credentials.json', 
                              token_path: str = 'token.json') -> Dict[str, Any]:
        """
        Save results to a beautifully formatted Google Spreadsheet
        
        Args:
            title: Title for the Google Spreadsheet
            credentials_path: Path to credentials.json
            token_path: Path to token.json
            
        Returns:
            Dictionary with spreadsheet_id and spreadsheet_url
        """
        # Initialize Google Sheets Manager
        sheets_manager = GoogleSheetsManager(credentials_path, token_path)
        
        # Create spreadsheet
        result = sheets_manager.create_spreadsheet(title)
        spreadsheet_id = result['spreadsheet_id']
        
        # Get DataFrame
        df = self.generate_dataframe()
        
        # Convert DataFrame to list of lists (including header)
        data = [df.columns.tolist()] + df.values.tolist()
        
        # Write data to default sheet (Sheet1)
        sheets_manager.write_data(spreadsheet_id, 'Sheet1', data)
        
        # Rename Sheet1 to "Sentences"
        sheet1_id = sheets_manager.get_sheet_id(spreadsheet_id, 'Sheet1')
        if sheet1_id is not None:
            rename_request = [{
                'updateSheetProperties': {
                    'properties': {
                        'sheetId': sheet1_id,
                        'title': 'Sentences'
                    },
                    'fields': 'title'
                }
            }]
            sheets_manager.format_sheet(spreadsheet_id, sheet1_id, rename_request)
            
            # Apply header formatting (blue background)
            sheets_manager.apply_header_formatting(
                spreadsheet_id, 
                sheet1_id, 
                len(df.columns),
                {'red': 0.27, 'green': 0.45, 'blue': 0.77}  # Blue #4472C4
            )
            
            # Apply borders
            sheets_manager.apply_borders(spreadsheet_id, sheet1_id, len(df) + 1, len(df.columns))
            
            # Freeze header row
            sheets_manager.freeze_rows(spreadsheet_id, sheet1_id, 1)
            
            # Set column widths (in pixels)
            column_widths_map = {
                0: 60,   # Row
                1: 500,  # Sentence
            }
            if 'Original' in df.columns:
                column_widths_map[2] = 500  # Original
                column_widths_map[3] = 150  # Method
                column_widths_map[4] = 100  # Word_Count
            else:
                column_widths_map[2] = 100  # Word_Count
            
            sheets_manager.set_column_widths(spreadsheet_id, sheet1_id, column_widths_map)
            
            # Color code rows based on method
            if 'Method' in df.columns:
                method_col_idx = df.columns.tolist().index('Method')
                color_requests = []
                
                for row_idx, row in enumerate(df.values, start=1):  # Start from 1 (after header)
                    method = row[method_col_idx]
                    if method:
                        if 'AI-Rewritten' in str(method):
                            # Light green
                            color = {'red': 0.91, 'green': 0.96, 'blue': 0.91}
                        elif 'Mechanical' in str(method):
                            # Light orange
                            color = {'red': 1.0, 'green': 0.96, 'blue': 0.90}
                        else:
                            # White (default)
                            color = {'red': 1.0, 'green': 1.0, 'blue': 1.0}
                        
                        color_requests.append({
                            'repeatCell': {
                                'range': {
                                    'sheetId': sheet1_id,
                                    'startRowIndex': row_idx,
                                    'endRowIndex': row_idx + 1,
                                    'startColumnIndex': 0,
                                    'endColumnIndex': len(df.columns)
                                },
                                'cell': {
                                    'userEnteredFormat': {
                                        'backgroundColor': color
                                    }
                                },
                                'fields': 'userEnteredFormat.backgroundColor'
                            }
                        })
                
                # Apply colors in batches
                if color_requests:
                    batch_size = 100
                    for i in range(0, len(color_requests), batch_size):
                        batch = color_requests[i:i + batch_size]
                        sheets_manager.format_sheet(spreadsheet_id, sheet1_id, batch)
        
        # Add Processing Log sheet if enabled
        if self.config.get_generate_log():
            log_df = self.generate_processing_log()
            if not log_df.empty:
                sheets_manager.create_sheet(spreadsheet_id, 'Processing Log', 
                                           row_count=len(log_df) + 10,
                                           column_count=len(log_df.columns))
                
                log_data = [log_df.columns.tolist()] + log_df.values.tolist()
                sheets_manager.write_data(spreadsheet_id, 'Processing Log', log_data)
                
                log_sheet_id = sheets_manager.get_sheet_id(spreadsheet_id, 'Processing Log')
                if log_sheet_id is not None:
                    # Apply header formatting (green background)
                    sheets_manager.apply_header_formatting(
                        spreadsheet_id,
                        log_sheet_id,
                        len(log_df.columns),
                        {'red': 0.44, 'green': 0.68, 'blue': 0.28}  # Green #70AD47
                    )
                    
                    # Apply borders
                    sheets_manager.apply_borders(spreadsheet_id, log_sheet_id, 
                                                len(log_df) + 1, len(log_df.columns))
                    
                    # Freeze header row
                    sheets_manager.freeze_rows(spreadsheet_id, log_sheet_id, 1)
                    
                    # Set column widths
                    log_widths = {i: 250 for i in range(len(log_df.columns))}
                    sheets_manager.set_column_widths(spreadsheet_id, log_sheet_id, log_widths)
        
        # Add Summary sheet
        summary = self.get_summary()
        summary_data = [
            ['Metric', 'Value'],
            ['Total Input Sentences', summary['total_input_sentences']],
            ['Total Output Sentences', summary['total_output_sentences']],
            ['Direct (No Processing)', summary['direct_sentences']],
            ['AI Rewritten', summary['ai_rewritten']],
            ['Mechanical Chunked', summary['mechanical_chunked']],
            ['Processing Time', f"{summary.get('processing_time', 0):.2f}s"],
            ['Average Words per Sentence', 
             f"{summary['total_output_sentences'] / summary['total_input_sentences']:.2f}" 
             if summary['total_input_sentences'] > 0 else 'N/A'],
            ['Success Rate', 
             f"{((summary['total_input_sentences'] - summary.get('failed', 0)) / summary['total_input_sentences'] * 100):.1f}%" 
             if summary['total_input_sentences'] > 0 else 'N/A']
        ]
        
        sheets_manager.create_sheet(spreadsheet_id, 'Summary', row_count=20, column_count=2)
        sheets_manager.write_data(spreadsheet_id, 'Summary', summary_data)
        
        summary_sheet_id = sheets_manager.get_sheet_id(spreadsheet_id, 'Summary')
        if summary_sheet_id is not None:
            # Apply header formatting (orange background)
            sheets_manager.apply_header_formatting(
                spreadsheet_id,
                summary_sheet_id,
                2,
                {'red': 0.93, 'green': 0.49, 'blue': 0.19}  # Orange #ED7D31
            )
            
            # Apply borders
            sheets_manager.apply_borders(spreadsheet_id, summary_sheet_id, len(summary_data), 2)
            
            # Freeze header row
            sheets_manager.freeze_rows(spreadsheet_id, summary_sheet_id, 1)
            
            # Set column widths
            sheets_manager.set_column_widths(spreadsheet_id, summary_sheet_id, {0: 250, 1: 150})
            
            # Bold metric names
            bold_requests = [{
                'repeatCell': {
                    'range': {
                        'sheetId': summary_sheet_id,
                        'startRowIndex': 1,
                        'endRowIndex': len(summary_data),
                        'startColumnIndex': 0,
                        'endColumnIndex': 1
                    },
                    'cell': {
                        'userEnteredFormat': {
                            'textFormat': {
                                'bold': True
                            }
                        }
                    },
                    'fields': 'userEnteredFormat.textFormat.bold'
                }
            }]
            sheets_manager.format_sheet(spreadsheet_id, summary_sheet_id, bold_requests)
        
        return result
    
    def get_summary(self) -> dict:
        """
        Get processing summary
        
        Returns:
            Dictionary with summary statistics
        """
        # Prefer live stats from a SentenceSplitter if available
        try:
            # If results were produced by a splitter instance, that splitter may have stats
            # attached via the ai_rewriter or internal counters. Try to use those if present.
            # Fallback to computing from self.results.
            total_sentences = len(self.results)
            direct = sum(1 for r in self.results if r.method == "Direct")
            ai_rewritten = sum(1 for r in self.results if "AI-Rewritten" in r.method)
            mechanical = sum(1 for r in self.results if "Mechanical" in r.method)
            failed = sum(1 for r in self.results if not r.success)
            total_output_sentences = sum(len(r.output_sentences) for r in self.results)

            summary = {
                'total_input_sentences': total_sentences,
                'total_output_sentences': total_output_sentences,
                'direct_sentences': direct,
                'ai_rewritten': ai_rewritten,
                'mechanical_chunked': mechanical,
                'failed': failed,
                'processing_time': self.processing_time,
                # Provide api_calls and cost if available from an AI rewriter
                'api_calls': 0,
                'cost': 0.0
            }

            # If any of the results include ai token stats, accumulate them
            # Search for ai_rewriter token stats on results (best-effort)
            for r in self.results:
                if hasattr(r, 'ai_token_cost'):
                    try:
                        summary['cost'] += float(getattr(r, 'ai_token_cost') or 0.0)
                    except Exception:
                        pass
            # Merge live stats from the active splitter if available
            try:
                splitter = getattr(self, '_active_splitter', None)
                if splitter is not None:
                    sstats = splitter.get_stats() if hasattr(splitter, 'get_stats') else getattr(splitter, 'stats', {})
                    if isinstance(sstats, dict):
                        summary['api_calls'] = sstats.get('api_calls', summary.get('api_calls', 0))
                        # Try a few possible keys for cost
                        summary['cost'] = sstats.get('cost', sstats.get('token_cost', summary.get('cost', 0.0)))
            except Exception:
                pass

            return summary

        except Exception as e:
            # Fallback simple summary
            return {
                'total_input_sentences': len(self.results),
                'total_output_sentences': sum(len(r.output_sentences) for r in self.results),
                'direct_sentences': sum(1 for r in self.results if r.method == "Direct"),
                'ai_rewritten': sum(1 for r in self.results if "AI-Rewritten" in r.method),
                'mechanical_chunked': sum(1 for r in self.results if "Mechanical" in r.method),
                'failed': sum(1 for r in self.results if not r.success),
                'processing_time': self.processing_time
            }
